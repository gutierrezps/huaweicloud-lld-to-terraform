import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet


def clean_str(input_str: str) -> str:
    """Transform into lowercase and replace non-alphanumeric with '_ '
    """
    ret = input_str.lower()
    ret = re.sub('[^0-9a-zA-Z]+', '_', ret)
    return ret


def cell_value(sheet: Worksheet, coord: str) -> Any:
    """Return cell value, whether it's merged or not.

    For example, if cells A1 and A2 are merged, the value is stored
    only in the first cell of the range (A1), and the other cells
    have an empty value.

    Using this function, the same value is returned for all cells in the
    merge range.

    Source: <https://stackoverflow.com/a/67610996/2014507>

    Args:
        sheet (Worksheet):
        coord (str): cell coordinate (e.g. 'A2')

    Raises:
        AssertionError: something is wrong with the spreadsheet

    Returns:
        Any: cell value, which can be date, str, int, float or None if
            it's empty
    """
    cell: Cell = sheet[coord]
    if not isinstance(cell, MergedCell):
        return cell.value

    # "Oh no, the cell is merged!"
    for range in sheet.merged_cells.ranges:
        if coord in range:
            return range.start_cell.value

    raise AssertionError('Merged cell is not in any merge range!')


def load_metadata(filename: str) -> dict:
    """Load parameters set on metadata.xlsx, which tells where to
    look for information in the LLD spreadsheet.

    Returns:
        dict: key is parameter name (cleaned up)
    """
    wb = load_workbook(filename)
    worksheet: Worksheet = wb['Parameters']

    params = {}

    for row in worksheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[0] is None:
            break

        param_name = clean_str(row[0])
        if param_name in params:
            print('[ERR] duplicate parameter:', param_name)
            exit(-1)

        params[param_name] = row[1]

    return params


def load_sheet_columns(worksheet: Worksheet) -> dict:
    """Read first row of worksheet, extract parameter names and
    associate to their respective column letters.

    Args:
        worksheet (Worksheet): input worksheet

    Returns:
        dict: key is parameter name (cleaned), value is column letter.
            E.g. if cell B1 has value ""
    """
    columns = {}

    for col in worksheet.iter_cols(min_col=1, min_row=1, max_row=1):
        header: Cell = col[0]
        if header.value is None:
            break
        param = clean_str(header.value.strip())
        if param in columns:
            msg = f'[ERR] Duplicate header in sheet {worksheet.title}: '
            msg += f'{param}'
            print(msg)
            exit(-1)

        columns[param] = header.column_letter

    return columns


def load_sheet_data(worksheet: Worksheet) -> list:
    """Load worksheet data, using header row as parameter names.

    Returns a list of dicts, where each dict is a worksheet row, and in
    each dict the key is parameter name (header) and value is the
    respective cell value.

    If a cell is empty for a given parameter, the parameter will not
    be present in that row data.

    The loading will stop at the first empty row.

    Args:
        worksheet (Worksheet): input worksheet

    Returns:
        list: of dicts, with data of each row
    """
    columns = load_sheet_columns(worksheet)

    sheet_data = {}

    for row in worksheet.iter_rows(min_row=2):
        row_index = row[0].row
        sheet_data[row_index] = {}
        row_is_empty = True

        for param, col in columns.items():
            value = cell_value(worksheet, col + str(row_index))
            if value is None:
                continue
            row_is_empty = False
            sheet_data[row_index][param] = value

        if row_is_empty:
            del sheet_data[row_index]
            break

    return sheet_data
